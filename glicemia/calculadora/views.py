from datetime import timedelta
from io import BytesIO

from django.contrib.auth.decorators import login_required, user_passes_test
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .forms import GlucemiaForm
from .models import MedicionGlucemia
from .services import evaluar_glucemia_service

# =========================================================
# PERMISOS
# =========================================================

def tiene_acceso_home(user):
    """
    Permite acceso solo a:
    - superusuarios
    - grupo Enfermeria
    - grupo Medicos
    """
    return user.is_authenticated and (
        user.is_superuser
        or user.groups.filter(name="Enfermeria").exists()
        or user.groups.filter(name="Medicos").exists()
    )

# =========================================================
# HELPERS DE VISTA
# =========================================================

def _normalizar_clase_desde_estado(estado):
    """
    Traduce el estado del resultado del service a una clase resumida
    para guardar en la BD y facilitar filtros.
    """
    if not estado:
        return "sin_clasificacion"

    estado = str(estado).lower()

    if "hipogluc" in estado:
        return "hipoglucemia"

    if (
        "post_hipo" in estado
        or "post_hipogluc" in estado
        or "rebote_post_hipoglucemia" in estado
    ):
        return "post_hipoglucemia"

    if (
        "rango" in estado
        or "objetivo" in estado
        or "estable en rango" in estado
        or "estable_en_rango" in estado
    ):
        return "en_rango"

    if (
        "hipergluc" in estado
        or "fuera de objetivo" in estado
        or "fuera_objetivo" in estado
    ):
        return "hiperglucemia"

    return "sin_clasificacion"


def _asignar_clase_visual(resultado):
    """
    Define la clase visual para la UI:
    - critico: hipo, hiper persistente/refractaria, alertas mayores
    - alerta: hiper aislada / fuera de objetivo no crítica
    - rango: en objetivo / en rango
    """
    if not resultado:
        return resultado

    estado = str(resultado.get("estado") or "").lower()
    subestado = str(resultado.get("subestado") or "").lower()

    if "hipogluc" in estado:
        resultado["clase_visual"] = "critico"
        return resultado

    if "persistente" in estado or "refractaria" in estado:
        resultado["clase_visual"] = "critico"
        return resultado

    if (
        "fuera de objetivo" in estado
        and "persistente" not in estado
        and "refractaria" not in estado
    ):
        resultado["clase_visual"] = "alerta"
        return resultado

    if "hipergluc" in estado:
        if "aislada" in estado or "aislada" in subestado:
            resultado["clase_visual"] = "alerta"
        else:
            resultado["clase_visual"] = "alerta"
        return resultado

    if (
        "rango" in estado
        or "objetivo" in estado
        or "estable en rango" in estado
        or "estable_en_rango" in estado
    ):
        resultado["clase_visual"] = "rango"
        return resultado

    resultado["clase_visual"] = "rango"
    return resultado


def _texto_seguro(valor):
    """
    Convierte None a string vacío para guardar o mostrar.
    """
    return "" if valor is None else str(valor)


def _guardar_medicion(request, cleaned_data, resultado):
    """
    Guarda en la base la medición calculada a partir del form y el resultado.
    """
    if not request.user.is_authenticated:
        return None

    estado = resultado.get("estado")
    clase = _normalizar_clase_desde_estado(estado)

    medicion = MedicionGlucemia.objects.create(
        usuario=request.user,
        glicemia_actual=int(cleaned_data["glicemia_actual"]),
        glicemia_previa=(
            int(cleaned_data["glicemia_previa"])
            if cleaned_data.get("glicemia_previa") is not None
            else None
        ),
        infusion_activa=bool(cleaned_data.get("infusion_activa")),
        hubo_ajuste_insulina=bool(cleaned_data.get("hubo_ajuste_insulina")),
        tercera_medicion=(
            int(cleaned_data["tercera_medicion"])
            if cleaned_data.get("tercera_medicion") is not None
            else None
        ),
        modo=cleaned_data.get("modo") or "seguimiento",
        estado=_texto_seguro(resultado.get("estado")),
        subestado=_texto_seguro(resultado.get("subestado")),
        clase=clase,
        mensaje=_texto_seguro(resultado.get("mensaje")),
        conducta=_texto_seguro(resultado.get("conducta")),
        proximo_control=_texto_seguro(resultado.get("proximo_control")),
        observacion=_texto_seguro(
            resultado.get("observacion")
            or resultado.get("comentario_control")
            or resultado.get("conducta_extra")
        ),
        tendencia=_texto_seguro(resultado.get("tendencia")),
        flecha_tendencia=_texto_seguro(resultado.get("flecha_tendencia")),
        delta=_texto_seguro(resultado.get("delta")),
        algoritmo_usado=_texto_seguro(
            resultado.get("algoritmo_usado") or resultado.get("algoritmo_sugerido")
        ),
        velocidad_sugerida=_texto_seguro(resultado.get("velocidad_sugerida")),
        bolo_ui=_texto_seguro(resultado.get("bolo_inicial")),
        tasa_inicial_ui_h=_texto_seguro(resultado.get("tasa_inicial")),
        tasa_algoritmo=_texto_seguro(resultado.get("tasa_algoritmo")),
        requiere_recontrol=bool(resultado.get("requiere_recontrol")),
        suspender_insulina=bool(resultado.get("suspender_insulina")),
        administrar_dextrosa=bool(resultado.get("administrar_dextrosa")),
        reiniciar_insulina=bool(resultado.get("reiniciar_insulina")),
        alerta_hgr=(
            clase == "hiperglucemia" and "refractaria" in _texto_seguro(estado).lower()
        ),
    )

    return medicion


def _filtrar_mediciones_desde_request(request):
    """
    Filtra mediciones por:
    - usuario
    - estado
    - clase
    - periodo (semanal / mensual)
    """
    usuario = request.GET.get("usuario", "").strip()
    estado = request.GET.get("estado", "").strip()
    clase = request.GET.get("clase", "").strip()
    periodo = request.GET.get("periodo", "").strip().lower()

    mediciones = (
        MedicionGlucemia.objects.select_related("usuario").all().order_by("-fecha_hora")
    )

    if usuario:
        mediciones = mediciones.filter(usuario__username=usuario)

    if estado:
        mediciones = mediciones.filter(estado=estado)

    if clase:
        mediciones = mediciones.filter(clase=clase)

    ahora = timezone.now()

    if periodo == "semanal":
        desde = ahora - timedelta(days=7)
        mediciones = mediciones.filter(fecha_hora__gte=desde)
    elif periodo == "mensual":
        desde = ahora - timedelta(days=30)
        mediciones = mediciones.filter(fecha_hora__gte=desde)

    return mediciones, usuario, estado, clase, periodo


def _asignar_resumen_objetivo(resultado, infusion_activa):
    """
    Define el título superior del resultado:
    - En rango objetivo
    - Fuera de rango objetivo
    y el texto del rango usado.
    """
    if not resultado:
        return resultado

    clase_visual = resultado.get("clase_visual")

    if clase_visual == "rango":
        resultado["resumen_objetivo"] = "En rango objetivo"
    else:
        resultado["resumen_objetivo"] = "Fuera de rango objetivo"

    # Texto visible del rango objetivo usado por la UI
    if infusion_activa:
        resultado["texto_rango_objetivo"] = "Paciente insulinizado: 140 a 200 mg/dL"
    else:
        resultado["texto_rango_objetivo"] = "Paciente no insulinizado: 70 a 180 mg/dL"

    return resultado

# =========================================================
# VISTA PRINCIPAL
# =========================================================

@login_required
@user_passes_test(tiene_acceso_home, login_url="/login/")
def control_glicemia(request):
    resultado = None
    medicion_guardada = None

    if request.method == "POST":
        form = GlucemiaForm(request.POST)

        if form.is_valid():
            actual = form.cleaned_data["glicemia_actual"]
            previa = form.cleaned_data.get("glicemia_previa")
            infusion_activa = form.cleaned_data.get("infusion_activa")
            hubo_ajuste_insulina = form.cleaned_data.get("hubo_ajuste_insulina")
            tercera_medicion = form.cleaned_data.get("tercera_medicion")

            resultado = evaluar_glucemia_service(
                actual=actual,
                previa=previa,
                infusion_activa=infusion_activa,
                hubo_ajuste_insulina=hubo_ajuste_insulina,
                tercera_medicion=tercera_medicion,
            )

            resultado = _asignar_clase_visual(resultado)
            resultado = _asignar_resumen_objetivo(resultado, infusion_activa)

            medicion_guardada = _guardar_medicion(request, form.cleaned_data, resultado)
    else:
        form = GlucemiaForm()

    return render(
        request,
        "calculadora/control_glicemia.html",
        {
            "form": form,
            "resultado": resultado,
            "medicion_guardada": medicion_guardada,
        },
    )

# =========================================================
# HISTORIAL
# =========================================================

@login_required
@user_passes_test(tiene_acceso_home, login_url="/login/")
def historial(request):
    (
        mediciones_qs,
        usuario_seleccionado,
        estado_seleccionado,
        clase_seleccionada,
        periodo,
    ) = _filtrar_mediciones_desde_request(request)

    total = mediciones_qs.count()
    hipoglucemias = mediciones_qs.filter(clase="hipoglucemia").count()
    post_hipoglucemias = mediciones_qs.filter(clase="post_hipoglucemia").count()
    en_rango = mediciones_qs.filter(clase="en_rango").count()
    hiperglucemias = mediciones_qs.filter(clase="hiperglucemia").count()

    paginator = Paginator(mediciones_qs, 10)
    page_number = request.GET.get("page")
    mediciones = paginator.get_page(page_number)

    usuarios = (
        MedicionGlucemia.objects.values_list("usuario__username", flat=True)
        .distinct()
        .order_by("usuario__username")
    )

    estados = (
        MedicionGlucemia.objects.values_list("estado", flat=True)
        .distinct()
        .order_by("estado")
    )

    clases = (
        MedicionGlucemia.objects.values_list("clase", flat=True)
        .distinct()
        .order_by("clase")
    )

    return render(
        request,
        "calculadora/historial.html",
        {
            "mediciones": mediciones,
            "usuarios": usuarios,
            "estados": estados,
            "clases": clases,
            "usuario_seleccionado": usuario_seleccionado,
            "estado_seleccionado": estado_seleccionado,
            "clase_seleccionada": clase_seleccionada,
            "periodo_seleccionado": periodo,
            "total": total,
            "hipoglucemias": hipoglucemias,
            "post_hipoglucemias": post_hipoglucemias,
            "en_rango": en_rango,
            "hiperglucemias": hiperglucemias,
        },
    )

# =========================================================
# EXPORTAR EXCEL
# =========================================================

@login_required
@user_passes_test(tiene_acceso_home, login_url="/login/")
def exportar_historial_excel(request):
    mediciones, usuario, estado, clase, periodo = _filtrar_mediciones_desde_request(
        request
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Historial"

    titulo = f"Reporte de mediciones - {periodo or 'completo'}"
    ws["A1"] = titulo
    ws["A1"].font = Font(bold=True, size=14)

    ws["A2"] = f"Usuario: {usuario or 'Todos'}"
    ws["B2"] = f"Estado: {estado or 'Todos'}"
    ws["C2"] = f"Clase: {clase or 'Todas'}"
    ws["D2"] = f"Generado: {timezone.localtime().strftime('%d/%m/%Y %H:%M')}"

    headers = [
        "Fecha",
        "Usuario",
        "Glucemia actual",
        "Glucemia previa",
        "Infusión activa",
        "Modo",
        "Estado",
        "Subestado",
        "Clase",
        "Conducta",
        "Próximo control",
        "Tendencia",
        "Flecha",
    ]

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    fila_header = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=fila_header, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    fila = 5
    for m in mediciones:
        ws.cell(
            row=fila,
            column=1,
            value=timezone.localtime(m.fecha_hora).strftime("%d/%m/%Y %H:%M"),
        )
        ws.cell(row=fila, column=2, value=m.usuario.username if m.usuario else "")
        ws.cell(row=fila, column=3, value=f"{m.glicemia_actual} mg/dL")
        ws.cell(
            row=fila,
            column=4,
            value=f"{m.glicemia_previa} mg/dL" if m.glicemia_previa is not None else "",
        )
        ws.cell(row=fila, column=5, value="Sí" if m.infusion_activa else "No")
        ws.cell(row=fila, column=6, value=m.get_modo_display())
        ws.cell(row=fila, column=7, value=m.estado)
        ws.cell(row=fila, column=8, value=m.subestado)
        ws.cell(row=fila, column=9, value=m.clase)
        ws.cell(row=fila, column=10, value=m.conducta)
        ws.cell(row=fila, column=11, value=m.proximo_control)
        ws.cell(row=fila, column=12, value=m.tendencia or "")
        ws.cell(row=fila, column=13, value=m.flecha_tendencia or "")
        fila += 1

    widths = {
        "A": 18,
        "B": 18,
        "C": 16,
        "D": 16,
        "E": 14,
        "F": 18,
        "G": 24,
        "H": 30,
        "I": 18,
        "J": 38,
        "K": 24,
        "L": 18,
        "M": 10,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    nombre = f"historial_{periodo or 'completo'}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{nombre}"'
    return response


# =========================================================
# EXPORTAR PDF
# =========================================================


@login_required
@user_passes_test(tiene_acceso_home, login_url="/login/")
def exportar_historial_pdf(request):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (Paragraph, SimpleDocTemplate, Spacer,
                                    Table, TableStyle)

    mediciones, usuario, estado, clase, periodo = _filtrar_mediciones_desde_request(
        request
    )

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
    )

    styles = getSampleStyleSheet()
    elements = []

    elements.append(
        Paragraph(f"Reporte de mediciones - {periodo or 'completo'}", styles["Title"])
    )
    elements.append(Spacer(1, 8))
    elements.append(Paragraph(f"Usuario: {usuario or 'Todos'}", styles["Normal"]))
    elements.append(Paragraph(f"Estado: {estado or 'Todos'}", styles["Normal"]))
    elements.append(Paragraph(f"Clase: {clase or 'Todas'}", styles["Normal"]))
    elements.append(
        Paragraph(
            f"Generado: {timezone.localtime().strftime('%d/%m/%Y %H:%M')}",
            styles["Normal"],
        )
    )
    elements.append(Spacer(1, 12))

    data = [
        [
            "Fecha",
            "Usuario",
            "Actual",
            "Previa",
            "Infusión",
            "Estado",
            "Clase",
            "Conducta",
            "Tendencia",
        ]
    ]

    for m in mediciones:
        data.append(
            [
                timezone.localtime(m.fecha_hora).strftime("%d/%m/%Y %H:%M"),
                m.usuario.username if m.usuario else "",
                f"{m.glicemia_actual}",
                f"{m.glicemia_previa}" if m.glicemia_previa is not None else "",
                "Sí" if m.infusion_activa else "No",
                m.estado,
                m.clase,
                m.conducta,
                f"{m.tendencia or ''} {m.flecha_tendencia or ''}".strip(),
            ]
        )

    table = Table(
        data,
        colWidths=[
            28 * mm,
            28 * mm,
            18 * mm,
            18 * mm,
            18 * mm,
            34 * mm,
            24 * mm,
            60 * mm,
            24 * mm,
        ],
    )

    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#C7D3E0")),
                ("BACKGROUND", (0, 1), (-1, -1), colors.whitesmoke),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                ("TOPPADDING", (0, 0), (-1, 0), 8),
            ]
        )
    )

    elements.append(table)
    doc.build(elements)

    pdf = buffer.getvalue()
    buffer.close()

    nombre = f"historial_{periodo or 'completo'}.pdf"
    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{nombre}"'
    return response
