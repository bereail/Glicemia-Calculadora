import json
from datetime import timedelta
from decimal import Decimal
from io import BytesIO

from django.contrib.auth.decorators import login_required, user_passes_test
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .forms import GlucemiaForm
from .models import MedicionGlucemia
from .services import evaluar_glucemia_service
from .utils.helpers import _a_bool
from .utils.reportes.metricas import calcular_detalle_categoria, calcular_metricas_dashboard
from .utils.reportes.metricas_pdf import generar_pdf_metricas
from .utils.reportes.registro_pdf import generar_pdf_registro
from .utils.ui.presentation import (
    enriquecer_resultado_ui,
    normalizar_clase_desde_estado,
    texto_seguro,
)


# =========================================================
# PERMISOS
# =========================================================

def _calcular_contexto_infusion(user):
    primera = (
        MedicionGlucemia.objects
        .filter(usuario=user, infusion_activa=True)
        .order_by("fecha_hora")
        .first()
    )
    horas_desde_inicio = None
    if primera:
        delta = timezone.now() - primera.fecha_hora
        horas_desde_inicio = delta.total_seconds() / 3600

    ultimas = list(
        MedicionGlucemia.objects
        .filter(usuario=user, infusion_activa=True)
        .order_by("-fecha_hora")[:2]
    )
    estable = len(ultimas) >= 2 and all(m.clase == "en_rango" for m in ultimas)

    return horas_desde_inicio, estable


def tiene_acceso_home(user):
    return user.is_authenticated and (
        user.is_superuser
        or user.groups.filter(name="Enfermeria").exists()
        or user.groups.filter(name="Medicos").exists()
    )


def tiene_acceso_historial(user):
    return user.is_authenticated and (
        user.is_superuser
        or user.username == "metanutric"
        or user.groups.filter(name="Historial").exists()
    )


# =========================================================
# GUARDADO
# =========================================================

def _guardar_medicion(request, cleaned_data, resultado):
    if not request.user.is_authenticated:
        return None

    estado = resultado.get("estado")
    clase = normalizar_clase_desde_estado(estado)

    medicion = MedicionGlucemia.objects.create(
        usuario=request.user,
        glicemia_actual=int(cleaned_data["glicemia_actual"]),
        glicemia_previa=(
            int(cleaned_data["glicemia_previa"])
            if cleaned_data.get("glicemia_previa") is not None
            else None
        ),
        infusion_activa=_a_bool(cleaned_data.get("infusion_activa")),
        tercera_medicion=(
            int(cleaned_data["tercera_medicion"])
            if cleaned_data.get("tercera_medicion") is not None
            else None
        ),
        modo=cleaned_data.get("modo") or "seguimiento",
        estado=texto_seguro(resultado.get("estado")),
        subestado=texto_seguro(resultado.get("subestado")),
        clase=clase,
        mensaje=texto_seguro(resultado.get("mensaje")),
        conducta=texto_seguro(resultado.get("conducta")),
        proximo_control=texto_seguro(resultado.get("proximo_control")),
        observacion=texto_seguro(
            resultado.get("observacion")
            or resultado.get("conducta_extra")
        ),
        tendencia=texto_seguro(resultado.get("tendencia")),
        flecha_tendencia=texto_seguro(resultado.get("flecha_tendencia")),
        delta=texto_seguro(resultado.get("delta")),
        algoritmo_usado=texto_seguro(
            resultado.get("algoritmo_usado") or resultado.get("algoritmo_sugerido")
        ),
        velocidad_sugerida=texto_seguro(resultado.get("velocidad_sugerida")),
        bolo_ui=texto_seguro(resultado.get("bolo_inicial")),
        tasa_inicial_ui_h=texto_seguro(resultado.get("tasa_inicial")),
        tasa_algoritmo=texto_seguro(resultado.get("tasa_algoritmo")),
        requiere_recontrol=bool(resultado.get("requiere_recontrol")),
        suspender_insulina=bool(resultado.get("suspender_insulina")),
        administrar_dextrosa=bool(resultado.get("administrar_dextrosa")),
        reiniciar_insulina=bool(resultado.get("reiniciar_insulina")),
        alerta_hgr=(
            clase == "hiperglucemia"
            and "refractaria" in texto_seguro(estado).lower()
        ),
    )

    return medicion


# =========================================================
# VIEW PRINCIPAL
def _json_safe(obj):
    if isinstance(obj, dict):
        return {k: _json_safe(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [_json_safe(v) for v in obj]
    if isinstance(obj, Decimal):
        return float(obj)
    return obj


# =========================================================

@login_required
@user_passes_test(tiene_acceso_home, login_url="/login/")
def control_glicemia(request):
    if request.method == "POST":
        form = GlucemiaForm(request.POST)

        if form.is_valid():
            actual = form.cleaned_data["glicemia_actual"]
            previa = form.cleaned_data.get("glicemia_previa")
            infusion_activa = form.cleaned_data.get("infusion_activa")
            tercera_medicion = form.cleaned_data.get("tercera_medicion")
            algoritmo_activo = form.cleaned_data.get("algoritmo_activo", "1")

            horas_desde_inicio, estable = _calcular_contexto_infusion(request.user)

            resultado = evaluar_glucemia_service(
                actual=actual,
                previa=previa,
                infusion_activa=infusion_activa,
                tercera_medicion=tercera_medicion,
                algoritmo_activo=algoritmo_activo,
                horas_desde_inicio=horas_desde_inicio,
                estable=estable,
            )

            resultado = enriquecer_resultado_ui(
                resultado=resultado,
                actual=actual,
                infusion_activa=_a_bool(infusion_activa),
            )

            resultado["infusion_activa"] = _a_bool(infusion_activa)

            estado_lower = str(resultado.get("estado") or "").lower()

            if "persistente" in estado_lower or "refractaria" in estado_lower:
                resultado["es_critico"] = True
                resultado["nivel_visual"] = "critico"
                resultado["clase_visual"] = "alerta"
            elif "hipergluc" in estado_lower:
                resultado["clase_visual"] = "alerta"

            _guardar_medicion(request, form.cleaned_data, resultado)

            request.session["glicemia_resultado"] = _json_safe(resultado)
            request.session["glicemia_post_data"] = request.POST.dict()
            return redirect("control_glicemia")

        return render(
            request,
            "calculadora/control_glicemia.html",
            {
                "form": form,
                "resultado": None,
                "medicion_guardada": None,
                "es_medico": request.user.groups.filter(name="Medicos").exists(),
            },
        )

    resultado = request.session.pop("glicemia_resultado", None)
    post_data = request.session.pop("glicemia_post_data", None)

    if post_data:
        form = GlucemiaForm(post_data)
        form.is_valid()
    else:
        form = GlucemiaForm()

    return render(
        request,
        "calculadora/control_glicemia.html",
        {
            "form": form,
            "resultado": resultado,
            "medicion_guardada": resultado is not None,
            "es_medico": request.user.groups.filter(name="Medicos").exists(),
        },
    )



_TURNOS = {
    "manana":    (6, 12),
    "tarde":     (12, 18),
    "noche":     (18, 24),
    "madrugada": (0, 6),
}


def _filtrar_mediciones_desde_request(request):
    from datetime import datetime as dt
    usuario = request.GET.get("usuario", "").strip()
    estado = request.GET.get("estado", "").strip()
    clase = request.GET.get("clase", "").strip()
    periodo = request.GET.get("periodo", "").strip().lower()
    turno = request.GET.get("turno", "").strip().lower()
    fecha_desde_str = request.GET.get("fecha_desde", "").strip()
    fecha_hasta_str = request.GET.get("fecha_hasta", "").strip()

    mediciones = (
        MedicionGlucemia.objects.select_related("usuario")
        .all()
        .order_by("-fecha_hora")
    )

    if usuario:
        mediciones = mediciones.filter(usuario__username=usuario)

    if estado:
        mediciones = mediciones.filter(estado=estado)

    if clase:
        mediciones = mediciones.filter(clase=clase)

    # Filtro por rango de fechas (tiene precedencia sobre período)
    if fecha_desde_str:
        try:
            desde = timezone.make_aware(dt.strptime(fecha_desde_str, "%Y-%m-%d"))
            mediciones = mediciones.filter(fecha_hora__gte=desde)
        except ValueError:
            fecha_desde_str = ""

    if fecha_hasta_str:
        try:
            hasta = timezone.make_aware(
                dt.strptime(fecha_hasta_str, "%Y-%m-%d") + timedelta(days=1)
            )
            mediciones = mediciones.filter(fecha_hora__lt=hasta)
        except ValueError:
            fecha_hasta_str = ""

    # Período preestablecido (solo si no hay fecha manual)
    if not fecha_desde_str and not fecha_hasta_str:
        ahora = timezone.now()
        if periodo == "semanal":
            mediciones = mediciones.filter(fecha_hora__gte=ahora - timedelta(days=7))
        elif periodo == "mensual":
            mediciones = mediciones.filter(fecha_hora__gte=ahora - timedelta(days=30))

    if turno and turno in _TURNOS:
        h_ini, h_fin = _TURNOS[turno]
        ids = [
            pk for pk, fecha in mediciones.values_list("pk", "fecha_hora")
            if h_ini <= timezone.localtime(fecha).hour < h_fin
        ]
        mediciones = mediciones.filter(pk__in=ids)

    return mediciones, usuario, estado, clase, periodo, turno, fecha_desde_str, fecha_hasta_str


@login_required
@user_passes_test(tiene_acceso_historial, login_url="/login/")
def historial(request):
    (
        mediciones_qs,
        usuario_seleccionado,
        estado_seleccionado,
        clase_seleccionada,
        periodo,
        turno_seleccionado,
        fecha_desde_seleccionada,
        fecha_hasta_seleccionada,
    ) = _filtrar_mediciones_desde_request(request)

    metricas = calcular_metricas_dashboard(mediciones_qs)

    paginator = Paginator(mediciones_qs, 5)
    page_number = request.GET.get("page")
    mediciones = paginator.get_page(page_number)

    usuarios = (
        MedicionGlucemia.objects.values_list("usuario__username", flat=True)
        .distinct()
        .order_by("usuario__username")
    )

    estados = (
        MedicionGlucemia.objects.values_list("estado", flat=True)
        .distinct()
        .exclude(estado="Hiperglucemia Severa")
        .order_by("estado")
    )

    clases = (
        MedicionGlucemia.objects.values_list("clase", flat=True)
        .distinct()
        .order_by("clase")
    )

    contexto = {
        "mediciones": mediciones,
        "usuarios": usuarios,
        "estados": estados,
        "clases": clases,
        "usuario_seleccionado": usuario_seleccionado,
        "estado_seleccionado": estado_seleccionado,
        "clase_seleccionada": clase_seleccionada,
        "periodo_seleccionado": periodo,
        "turno_seleccionado": turno_seleccionado,
        "fecha_desde_seleccionada": fecha_desde_seleccionada,
        "fecha_hasta_seleccionada": fecha_hasta_seleccionada,
        **metricas,
    }
    return render(request, "calculadora/historial.html", contexto)


@login_required
@user_passes_test(tiene_acceso_historial, login_url="/login/")
def historial_detalle(request):
    """Detalle de una categoría (card) del panel, respetando los mismos filtros
    (fecha/turno/usuario/estado) que están aplicados en la página de historial."""
    clase = request.GET.get("clase", "").strip()
    LIMITE = 200

    qs, *_ = _filtrar_mediciones_desde_request(request)

    total_clase = qs.count()
    mediciones = list(qs[:LIMITE])
    detalle_graficos = calcular_detalle_categoria(mediciones)

    contexto = {
        "mediciones": mediciones,
        "total_clase": total_clase,
        "hay_mas": total_clase > LIMITE,
        "clase": clase,
        **detalle_graficos,
    }
    return render(request, "calculadora/partials/detalle_clase.html", contexto)


@login_required
@user_passes_test(tiene_acceso_historial, login_url="/login/")
def exportar_historial_excel(request):
    mediciones, usuario, estado, clase, periodo, turno, *_ = _filtrar_mediciones_desde_request(
        request
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Historial"

    titulo = f"Reporte de mediciones - {periodo or 'completo'}"
    ws["A1"] = titulo
    ws["A1"].font = Font(bold=True, size=14)

    ws["A2"] = f"Usuario: {usuario or 'Todos'}"
    ws["B2"] = f"Estado: {estado or 'Todos'}"
    ws["C2"] = f"Clase: {clase or 'Todas'}"
    ws["D2"] = f"Generado: {timezone.localtime().strftime('%d/%m/%Y %H:%M')}"

    headers = [
        "Fecha",
        "Usuario",
        "Glucemia actual",
        "Glucemia previa",
        "Infusión activa",
        "Modo",
        "Estado",
        "Subestado",
        "Clase",
        "Conducta",
        "Próximo control",
        "Tendencia",
        "Flecha",
    ]

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    fila_header = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=fila_header, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    fill_par = PatternFill("solid", fgColor="EBF3FB")
    fill_impar = PatternFill("solid", fgColor="FFFFFF")
    wrap = Alignment(wrap_text=True, vertical="top")
    center_top = Alignment(horizontal="center", vertical="top")

    fila = 5
    for m in mediciones:
        fill = fill_par if fila % 2 == 0 else fill_impar
        valores = [
            timezone.localtime(m.fecha_hora).strftime("%d/%m/%Y %H:%M"),
            m.usuario.username if m.usuario else "",
            f"{m.glicemia_actual} mg/dL",
            f"{m.glicemia_previa} mg/dL" if m.glicemia_previa is not None else "",
            "Sí" if m.infusion_activa else "No",
            m.get_modo_display(),
            m.estado,
            m.subestado,
            m.clase,
            (m.conducta or "").replace("<br>", " / ").replace("<br/>", " / "),
            m.proximo_control,
            m.tendencia or "",
            m.flecha_tendencia or "",
        ]
        for col, valor in enumerate(valores, start=1):
            cell = ws.cell(row=fila, column=col, value=valor)
            cell.fill = fill
            cell.alignment = wrap if col == 10 else center_top
        ws.row_dimensions[fila].height = 28
        fila += 1

    widths = {
        "A": 18, "B": 18, "C": 16, "D": 16, "E": 14,
        "F": 18, "G": 26, "H": 30, "I": 18, "J": 42,
        "K": 26, "L": 18, "M": 10,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    nombre = f"historial_{periodo or 'completo'}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{nombre}"'
    return response


_TURNO_TEXTOS = {
    "manana": "Mañana (06–12)",
    "tarde": "Tarde (12–18)",
    "noche": "Noche (18–24)",
    "madrugada": "Madrugada (00–06)",
}


def _construir_filtros_pdf(usuario, estado, clase, periodo, turno, fecha_desde, fecha_hasta):
    if fecha_desde or fecha_hasta:
        periodo_texto = f"{fecha_desde or '...'} – {fecha_hasta or '...'}"
    elif periodo == "semanal":
        periodo_texto = "Últimos 7 días"
    elif periodo == "mensual":
        periodo_texto = "Últimos 30 días"
    else:
        periodo_texto = "Completo"

    return {
        "periodo_texto": periodo_texto,
        "turno_texto": _TURNO_TEXTOS.get(turno),
        "usuario": usuario,
        "estado": estado,
        "clase": clase,
        "generado": timezone.localtime().strftime("%d/%m/%Y %H:%M"),
    }


def _nombre_archivo_filtrado(prefijo, periodo, turno, fecha_desde, fecha_hasta):
    partes = [prefijo]
    if fecha_desde:
        partes.append(fecha_desde)
    if fecha_hasta:
        partes.append(fecha_hasta)
    if turno:
        partes.append(turno)
    if not fecha_desde and not fecha_hasta and not turno:
        partes.append(periodo or "completo")
    return "_".join(partes) + ".pdf"


@login_required
@user_passes_test(tiene_acceso_historial, login_url="/login/")
def exportar_historial_pdf(request):
    (
        mediciones,
        usuario,
        estado,
        clase,
        periodo,
        turno,
        fecha_desde,
        fecha_hasta,
    ) = _filtrar_mediciones_desde_request(request)

    filtros = _construir_filtros_pdf(usuario, estado, clase, periodo, turno, fecha_desde, fecha_hasta)
    pdf = generar_pdf_registro(mediciones, filtros)

    nombre = _nombre_archivo_filtrado("historial", periodo, turno, fecha_desde, fecha_hasta)
    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{nombre}"'
    return response


@login_required
@user_passes_test(tiene_acceso_historial, login_url="/login/")
def exportar_metricas_pdf(request):
    (
        mediciones_qs,
        usuario,
        estado,
        clase,
        periodo,
        turno,
        fecha_desde,
        fecha_hasta,
    ) = _filtrar_mediciones_desde_request(request)

    metricas = calcular_metricas_dashboard(mediciones_qs)
    filtros = _construir_filtros_pdf(usuario, estado, clase, periodo, turno, fecha_desde, fecha_hasta)
    pdf = generar_pdf_metricas(metricas, filtros)

    nombre = _nombre_archivo_filtrado("metricas", periodo, turno, fecha_desde, fecha_hasta)
    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{nombre}"'
    return response